"""
This program performs the task of putting all the data of a particular candidate from 5 sheets
to a master sheet provided the user gives the input of name , ps no or email id of that particular candidate.
In master sheet we have to plot the bar graph of any row/column.

"""
# -------------------------------------------------------------------------------------------------------------#
# -------------------------------------------------------------------------------------------------------------#
"""
This program uses openpyxl library 

"""
import openpyxl
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference

# I am  taking the empty workbook
excel_file = Workbook()
excel_file = Workbook()
# loading my excel file
wb = openpyxl.load_workbook('studentinfo.xlsx')
# taking all the sheets in excel as the list
sheets = ['Sheet1', 'Sheet2', 'Sheet3', 'Sheet4', 'Sheet5']
# creating one mastersheet
excel_sheet = excel_file.create_sheet(title='MasterSheet11', index=0)
# taking input from the user like how many members information u want
n = int(input("Number of persons: "))
# loop for the number of persons
for g in range(1, n+1):
    print("--------Enter", g, " person information:--------")
    xin = int(input("Enter PS number: "))
    yin = input("Enter name: ")
    zin = input("Enter email: ")
# here i am taking my output in vertical(column) format
# here t is a variable
    t = 1
    for sheet in sheets:
        sh = wb[sheet]  # Get a sheet from the workbook.
        max_r = sh.max_row
        max_c = sh.max_column
        if t <= 10:     # for sheet1 this loop should be excecuted
            for r in range(1, max_r + 1):
                if sh.cell(row=r, column=1).value == xin and sh.cell(row=r, column=2).value == yin and sh.cell(row=r, column=3).value == zin:
                    print("Person data added.")
                    for c in range(1, max_c + 1):
                        if g == 1:    # for first person
                            str1 = 'A' + str(t)
                            str2 = 'B' + str(t)
                            t = t + 1
                            excel_sheet[str1] = str(sh.cell(row=1, column=c).value)
                            excel_sheet[str2] = sh.cell(row=r, column=c).value
                        else:
                            str1 = chr(67+g) + str(t)
                            str2 = chr(68+g) + str(t)
                            t = t + 1
                            excel_sheet[str1] = str(sh.cell(row=1, column=c).value)
                            excel_sheet[str2] = sh.cell(row=r, column=c).value
        # from sheet2 onwards this loop will be excecuted
        # because we dont want to repeat the name ps num and mail
        else:
            for r in range(4, max_r + 1):
                if sh.cell(row=r, column=1).value == xin and sh.cell(row=r, column=2).value == yin and sh.cell(row=r,
                                                                                                               column=3).value == zin:
                    for c in range(4, max_c + 1):
                        if g == 1:
                            str1 = 'A' + str(t)
                            str2 = 'B' + str(t)
                            t = t + 1
                            excel_sheet[str1] = str(sh.cell(row=1, column=c).value)
                            excel_sheet[str2] = sh.cell(row=r, column=c).value
                        else:
                            str1 = chr(67+g) + str(t)
                            str2 = chr(68+g) + str(t)
                            t = t + 1
                            excel_sheet[str1] = str(sh.cell(row=1, column=c).value)
                            excel_sheet[str2] = sh.cell(row=r, column=c).value
# output will be save in final.xlsx file
    excel_file.save(filename="final.xlsx")

# for ploting bar graph
wb = openpyxl.load_workbook('final.xlsx')
excel_file = wb
# Get workbook active sheet
# from the active attribute.
sheet = wb['MasterSheet11']
# create data for plotting like number of rows and columns data we want
values = Reference(sheet, min_col=2, min_row=9, max_col=sheet.max_column, max_row=17)
# Create object of BarChart class
chart = BarChart()
# adding data to the Bar chart object
chart.add_data(values)
# set the title of the chart
chart.title = " Student_Data "
# set the title of the x-axis
chart.x_axis.title = " Different_subjects "
# set the title of the y-axis
chart.y_axis.title = " Marks_scored "
# add chart to the sheet
# the top-left corner of a chart
# is anchored to cell E2 .
sheet.add_chart(chart, "J3")
# save the file
wb.save("final.xlsx")

print("Completed.")
